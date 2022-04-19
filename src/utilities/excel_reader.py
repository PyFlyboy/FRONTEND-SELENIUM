import openpyxl

path = "..\\report_test.xlsx"
workbook = openpyxl.load_workbook(path)
sheetName="sheet1"
sheet = workbook["report"]
totalrows = sheet.max_row
totalcols = sheet.max_column

def loadWorkbook(path):
    workbook = openpyxl.load_workbook(path)
    return workbook

def getRowCount(sheetName):
    wokrbook = loadWorkbook(path)
    sheet = wokrbook[sheetName]
    return sheet.max_row


def getColCount(sheetName):
    workbook = loadWorkbook(path)
    sheet = workbook[sheetName]
    return  sheet.max_column

def getCellData(sheetName, rowNum, colNum):
    workbook = loadWorkbook(path)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowNum, column=colNum)
    val = data.value
    return val

def setCellData(sheetName, rowNum, colNum, newData):
    workbook = loadWorkbook(path)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowNum, column=colNum)
    data.value = newData
    workbook.save(path)


